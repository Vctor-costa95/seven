import os
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def omissoes():
    path = 'C:\\Users\\jvict\\Documents\\seven\\downloads_malha\\malhas'
    files = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]

    lastCNPJ = ''
    arrayCNPJ = []
    arrayData = []
    arrayAnos = []

    for file in files:
        inputFileName = os.path.join(path, file)

        wb = openpyxl.load_workbook(inputFileName)
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            filterRow = list(filter(None, row))

            if filterRow and filterRow[0] == "CNPJ:":
                lastCNPJ = filterRow[1]
                arrayCNPJ.append(filterRow[1])

            if filterRow and len(filterRow[0]) == 44:
                chaveAcesso = separarChaveAcesso(filterRow[0])
                ano = chaveAcesso['AA']
                item = {
                    'cnpj': lastCNPJ,
                    'chave': chaveAcesso['chaveAcesso'],
                    'ano': chaveAcesso['AA'],
                    'tipoOperacao': filterRow[1] if len(filterRow) > 1 else '',
                    'valorNota': filterRow[2] if len(filterRow) > 2 else '0',
                    'multaIntegral': filterRow[3] if len(filterRow) > 3 else '0',
                    'multaEspontanea': filterRow[4] if len(filterRow) > 4 else '0'
                }

                arrayData.append(item)
                arrayAnos.append(f"{lastCNPJ}_{ano}")

    arrayCNPJ = list(set(arrayCNPJ))
    print(arrayCNPJ)
    arrayAnos = list(set(arrayAnos))
    print(arrayAnos)

    for cnpj in arrayCNPJ:
        nomeArquivo = cnpj.replace('.', '').replace('/', '').replace('-', '')
        wb_out = Workbook()
        sheet_out = wb_out.active

        sheet_out['D7'] = "CNPJ:"
        sheet_out['F7'] = cnpj

        initialPosition = 11
        for nota in arrayData:
            if nota['cnpj'] == cnpj:
                colunaChave = f"C{initialPosition}"
                sheet_out[colunaChave] = nota['chave']

                colunaTipoOperacao = f"P{initialPosition}"
                sheet_out[colunaTipoOperacao] = nota['tipoOperacao']

                colunaTipoOperacao = f"T{initialPosition}"
                sheet_out[colunaTipoOperacao] = float(nota['valorNota'])

                initialPosition += 1

        colunaFim = f"B{initialPosition}"
        sheet_out[colunaFim] = "FIM"

        wb_out.save(f"{nomeArquivo}.xlsx")
        print(f"Planilha gerada: {nomeArquivo}.xlsx")

def separarChaveAcesso(chave):
    # Implemente a lógica para separar a chave de acesso aqui
    # Substitua o retorno abaixo pelo código apropriado
    return {'chaveAcesso': chave, 'AA': chave[-2:]}

omissoes()
